import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DETAILS = Path("data/douyin_details_merged.json")
OUT = Path("output/feishu_import")


HEADERS = [
    "序号",
    "作品ID",
    "作品链接",
    "发布时间",
    "时长(秒)",
    "标题/简介",
    "转写状态",
    "文字稿",
    "备注",
]


def fmt_time(ts):
    if not ts:
        return ""
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    items = json.loads(DETAILS.read_text(encoding="utf-8"))
    rows = []
    for item in items:
        status = item.get("transcript_status", "")
        note = ""
        if status == "no_play_url":
            note = "详情接口未返回可播放视频流，需在抖音页面人工核对/补充。"
        rows.append(
            [
                item.get("index"),
                item.get("aweme_id"),
                item.get("share_url"),
                fmt_time(item.get("create_time")),
                round((item.get("duration_ms") or 0) / 1000, 1) if item.get("duration_ms") else "",
                item.get("desc", ""),
                status,
                item.get("transcript", ""),
                note,
            ]
        )

    csv_path = OUT / "抖音工作流_文字稿导入.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "抖音文字稿"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [8, 22, 36, 20, 10, 48, 14, 80, 36]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    xlsx_path = OUT / "抖音工作流_文字稿导入.xlsx"
    wb.save(xlsx_path)
    print(csv_path)
    print(xlsx_path)
    print(f"rows={len(rows)}")


if __name__ == "__main__":
    main()
